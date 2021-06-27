#Questão 4
#Bibliotecas necessárias importadas abaixo
#Aperte F5 para funcionar
#Comentário: ao rodar o código novamente, a planilha excel é reiniciada (fez mais sentido pra mim dessa forma). É possível alterar.


import pandas as pd

from openpyxl import Workbook

from openpyxl import load_workbook

'-------- Criação da classe banco de dados --------'

class Banco_dados:
    
    def __init__(self): 
        
        Bancodados_excel=Workbook()         #Criando um arquivo excel e salvando-o no mesmo local deste arquivo python 
        planilha1=Bancodados_excel.active
        planilha1.append(("DRE", "Curso", "Nome", "Genero", "Data_de_Nascimento", "Altura", "Peso", "CRA", "Creditos obtidos", "Renda"))
        Bancodados_excel.save("Bancodados.xlsx")

    def cadastro(self, DRE, Curso, Nome, Genero, Data_de_Nascimento, Altura, Peso, CRA, Creditos_obtidos, Renda):
        
        Bancodados_excel=load_workbook("Bancodados.xlsx")
        planilha1=Bancodados_excel.active
        linha=(DRE, Curso, Nome, Genero, Data_de_Nascimento, Altura, Peso, CRA, Creditos_obtidos, Renda)
        planilha1.append(linha)
        Bancodados_excel.save("Bancodados.xlsx")

    def estatistica(self,opcao): 
        
        df=pd.read_excel("Bancodados.xlsx")
        df.index
        
        if opcao == 1:

            cursoo=input("Qual curso desejado? \n")
            alunos=df.query("Curso == "+ "'" + cursoo +"'")

            if len(alunos.values) == 0:
            
                return "Nenhum aluno cadastrado neste curso"
            
            soma=0
            
            for cra in range(0,len(alunos.values)):
                
                soma+=alunos.loc[cra][7]
                
            return "A média de CRA do curso "+ cursoo+" é de " + str(soma/len(alunos.values))
        
        elif opcao == 2:
            
            soma=0
            
            for cra in range(0,len(df.values)):
                
                soma+=df.values[cra][7]
                
            if len(df.values)==0:
                
                return "Nenhum aluno cadastrado"
                
            return "A média de CRA geral é de "+str(soma/len(df.values))
        
        elif opcao == 3:
            
            curso = input("Qual curso desejado? \n")
            homens=df.query("Genero == 'M' & Curso == "+"'" + curso+"'")
            geral=df.query("Curso == "+ "'" + curso +"'")

            if len(geral.values)==0:
                
                return "Nenhum aluno cadastrado"
            
            return len(homens.values)/len(geral.values)
        
        elif opcao == 4:
            
            curso = input("Qual curso desejado? \n")
            mulheres=df.query("Genero == 'F' & Curso == "+"'" + curso+"'")
            geral=df.query("Curso == "+ "'" + curso +"'")

            if len(geral.values)==0:
                
                return "Nenhum aluno cadastrado"
            
            return len(mulheres.values)/len(geral.values)
        

'--------------- Criação da função responsável pelo loop interativo ----------'


def main():
    
    banco=Banco_dados()
    
    while True:
        
        resp=input("O que deseja fazer? \n \n 1. Cadastro de aluno \n 2. Consulta de dados \n Pressione s para fechar \n")
        
        if resp == "1":
            
            while True:
                
                a=input("DRE:\n")
                
                if type(int(a))== int and len(a)== 9:
                    
                    break
                
                else:
                    
                    print ("Formato inválido, DRE possui 9 dígitos numéricos, tente novamente:\n")

            b=input("Curso:\n")
            c=input("Nome:\n")
            
            while True:
                
                d=input("Genero (M ou F):\n")
                
                if d != "M" and d != "F":
                    
                    print("Favor inserir M ou F, tente novamente:\n")
                    
                else:
                    
                    break
                
            e=input("Data de nascimento (DD/MM/AAAA):\n")

            while True:
                
                f=input("Altura (em metros):\n")
                
                if type(float(f))== float and float(f) > 0:
                    
                    break
                
                else:
                    
                    print ("Altura inválida, tente novamente:\n")
                    
            while True:
                
                g=input("Peso (em quilos):\n")
                
                if type(float(g))== float and float(g) > 0:
                    
                    break
                
                else:
                    
                    print ("Peso inválido, tente novamente:\n")
                    
            while True:
                
                h=input("CRA:\n")
                
                if type(float(h))==float:
                    
                    break
                
                else:
                    
                    print ("CRA inválido, tente novamente:\n")
                    
            i=input("Creditos obtidos:\n")
            j=input("Renda (somente numero):\n")
            banco.cadastro(a,b,c,d,e,float(f),float(g),float(h),float(i),float(j))
            print ("\nAluno cadastrado com sucesso! \n")



        elif resp == "2":
            
            b=input("O que deseja consultar? \n \n 1. Média de CRA por curso \n 2. Média de CRA total \n 3. Proporção de alunos homem em um curso \n 4. Proporção de mulheres em um curso \n Pressione s para retornar ao menu anterior\n")
            
            if b=="1":

                print(banco.estatistica(int(b))) 
            
            elif b=="2":
                
                print(banco.estatistica(int(b)))
                
            elif b=="3":
                
                print(banco.estatistica(int(b)))
                
            elif b=="4":
                
                print(banco.estatistica(int(b)))
                
            else:
                
                print("Resposta inválida, retornando ao menu inicial\n")
                
        elif resp == "s":
            
            break


main()



